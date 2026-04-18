import sys
import Data.pytorch_autoencoder as pytorch_autoencoder
sys.modules["pytorch_autoencoder"] = pytorch_autoencoder
from Data.anomaly_detection import detect_anomalies, detect_using_isolation_forest, detect_using_one_class_support_vector_machine, add_synthetic_anomalies
from Data.autoencoder_data_preparation import read_and_prepare_data 
import tabulate
import os
import pandas as pd
import numpy as np
from collections import defaultdict
import random
import atexit
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

STORE = True
RANDOM_SEED = 42
random.seed(RANDOM_SEED)


def extract_event_column_indices(column_names):
    """Return indices and clean event names for one-hot encoded event columns."""
    event_columns = []
    for idx, name in enumerate(column_names):
        name_str = str(name)
        lower_name = name_str.lower()
        if "event" in lower_name:
            pos = lower_name.find("event")
            event_name = name_str[pos + len("event"):].strip(" _-:\t")
            if not event_name:
                event_name = name_str
            event_columns.append((idx, event_name))
    return event_columns


def get_event_for_row(row_values, event_columns):
    """Resolve the event name from one-hot encoded event values for one row."""
    if not event_columns:
        return "UNKNOWN_EVENT"

    active_events = [event_name for col_idx, event_name in event_columns if row_values[col_idx] > 0.5]
    if len(active_events) == 1:
        return active_events[0]

    # Fallback for non-binary/scaled values: use the event column with highest value.
    best_col_idx, best_event_name = max(event_columns, key=lambda col: row_values[col[0]])
    if row_values[best_col_idx] <= 0:
        return "UNKNOWN_EVENT"
    return best_event_name


def to_numpy_2d(array_like):
    """Convert tensor/array-like input to a 2D NumPy array for sklearn."""
    if hasattr(array_like, "detach"):
        array_like = array_like.detach()
    if hasattr(array_like, "cpu"):
        array_like = array_like.cpu()
    if hasattr(array_like, "numpy"):
        array_like = array_like.numpy()

    arr = np.asarray(array_like)
    if arr.ndim == 1:
        arr = arr.reshape(1, -1)
    elif arr.ndim > 2:
        arr = np.squeeze(arr)
        if arr.ndim != 2:
            raise ValueError(f"Expected 2D array for inverse scaling, got shape {arr.shape}")
    return arr

if __name__ == "__main__":
    tables_dir = "tables"
    if STORE:
        os.makedirs(tables_dir, exist_ok=True)
    
    # Ask for which resources anomaly detection should be performed
    resources = []
    all_resources = ["hbw_1", "hbw_2", "mm_1", "sm_1", "wt_2", "ov_1", "vgr_1", "ov_2", "mm_2", "dm_2", "pm_1", "wt_1", "sm_2", "vgr_2", "hw_1"]
    user_input = input("Do you want to perform anomaly detection on all resources (workstations)? (y/n):\n")
    if user_input.lower() == 'y':
        resources = all_resources
        print("\n")
    else:
        # Ask for individual resources
        print("Which resources do you want to perform anomaly detection on? Possible choices:")
        for resource in all_resources:
            print(" ", resource)
            
        print("For information about each resource, read README.")
        user_input = input("List all resources you want to use, seperated by commas:\n")
        resources = [resource.strip() for resource in user_input.split(',') if resource.strip() in all_resources]
        print("\n")
    
    # Ask, if the training data should be loaded from all the data again and then prepared
    user_input = input("Do you want to load and prepare training data from scratch? (y/n):\n")
    reread_prepared_data = not user_input.lower() == 'y'
    print("\n")
    
    # Ask, if the autoencoder should be retrained
    user_input = input("Do you want to retrain the autoencoder? (y/n):\n")
    retrain_AE = user_input.lower() == 'y'
    print("\n")
    
    # Ask if the hyperparameters should be redone
    redo_hyperparameter_tuning = False
    if retrain_AE:
        user_input = input("Do you want to redo the hyperparameter tuning? (y/n)\nWARNING: Redoing hyperparameter tuning takes a long time.\n")
        redo_hyperparameter_tuning = user_input.lower() == 'y'
        print("\n")
        
    all_results = {}
        
    for resource in resources:
        # Get the data
        print(f"Performing Anomaly detection on resource {resource}...\n\n" + ("Loading preprepared data...\n" if reread_prepared_data else "Reading and preparing data...\n"))
        train_scaled, test_scaled, val_scaled, scaling, column_names = read_and_prepare_data(resource, load_preprepared=reread_prepared_data, prints=False)
        
        val_anomalous, anomalous_values, anomaly_types = add_synthetic_anomalies(val_scaled, train_scaled, test_scaled, column_names)
        
        # Perform anomaly detection on the data
        print("Performing anomaly detection on the data...\n")
        detected_rows, reconstructions, threshold = detect_anomalies(train_scaled, test_scaled, val_anomalous, scaling, column_names, resource, train_model=retrain_AE, redo_hyperparameter_tuning=redo_hyperparameter_tuning, prints=True, val_scaled=val_scaled)
        
        # Perform anomaly detection on the data using baseline models for comparison
        print("Performing anomaly detection using baseline techniques...\n")
        if_detected_rows = detect_using_isolation_forest(train_scaled, test_scaled, val_anomalous, scaling, column_names, prints=True)
        svm_detected_rows = detect_using_one_class_support_vector_machine(train_scaled, val_anomalous, val_scaled, scaling, column_names, prints=True)
        
        all_results[resource] = {
            "train_scaled": train_scaled,
            "val_anomalous": val_anomalous,                 
            "anomalous_values": anomalous_values,       # Dict with row as id and the value is a numpy array containing the anomalous column-IDs    
            "ae_detected": detected_rows,               
            "if_detected": if_detected_rows,
            "svm_detected": svm_detected_rows,
            "column_names": column_names,
            "ae_reconstructions": reconstructions,                
            "scaling": scaling,
            "ae_threshold": threshold,
            "anomaly_types": anomaly_types
        }
        
    # Calculate precision, recall and F1 per resource
    methods = [("Autoencoder", "ae_detected"), 
               ("Isolation Forest", "if_detected"), 
               ("SVM", "svm_detected")]
    
    for method_name, method_key in methods:
        print(f"\n{'='*60}")
        print(f"Results for {method_name}")
        print(f"{'='*60}\n")
        
        results_table = []
        for resource, data in all_results.items():
            anomalous_values = data["anomalous_values"]
            train_rows = len(data["train_scaled"])
            val_rows = len(data["val_anomalous"])
            detected_rows = data[method_key]
            
            detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())
            tp = sum(1 for row_id in detected_set if row_id in anomalous_values)
            fp = sum(1 for row_id in detected_set if row_id not in anomalous_values)
            fn = sum(1 for row_id in anomalous_values.keys() if row_id not in detected_set)
            
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
            
            results_table.append([resource, train_rows, val_rows, f"{precision:.3f}", f"{recall:.3f}", f"{f1:.3f}"])
        
        results_headers = ["Resource", "Train Rows", "Val Rows", "Precision", "Recall", "F1-Score"]
        print(tabulate.tabulate(results_table, headers=results_headers))

        if STORE:
            # Store each per-method table in the tables folder.
            method_file_name = method_key.replace("_detected", "") + "_per_resource_metrics.xlsx"
            method_output_path = os.path.join(tables_dir, method_file_name)
            pd.DataFrame(results_table, columns=results_headers).to_excel(method_output_path, index=False)
        
    # Calculate combined results across all resources
    print(f"\n{'='*60}")
    print("Combined Results Across All Resources")
    print(f"{'='*60}\n")
    
    combined_table = []
    for method_name, method_key in methods:
        total_tp = 0
        total_fp = 0
        total_fn = 0
        
        for resource, data in all_results.items():
            anomalous_values = data["anomalous_values"]
            detected_rows = data[method_key]
            
            detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())
            total_tp += sum(1 for row_id in detected_set if row_id in anomalous_values)
            total_fp += sum(1 for row_id in detected_set if row_id not in anomalous_values)
            total_fn += sum(1 for row_id in anomalous_values.keys() if row_id not in detected_set)
        
        combined_precision = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0
        combined_recall = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0
        combined_f1 = 2 * (combined_precision * combined_recall) / (combined_precision + combined_recall) if (combined_precision + combined_recall) > 0 else 0
        
        combined_table.append([method_name, f"{combined_precision:.3f}", f"{combined_recall:.3f}", f"{combined_f1:.3f}"])
    
    combined_headers = ["Method", "Precision", "Recall", "F1-Score"]
    combined_table_text = tabulate.tabulate(combined_table, headers=combined_headers)
    print(combined_table_text)
    
    if STORE:
        combined_table_output_path = os.path.join(tables_dir, "combined_metrics.xlsx")
        pd.DataFrame(combined_table, columns=combined_headers).to_excel(combined_table_output_path, index=False)

    # Calculate metrics per anomaly type across all resources for all models
    print(f"\n{'='*60}")
    print("Results Per Anomaly Type")
    print(f"{'='*60}\n")

    for method_name, method_key in methods:
        print(f"\n{'-'*60}")
        print(f"{method_name}")
        print(f"{'-'*60}\n")

        anomaly_type_tp = defaultdict(int)
        anomaly_type_fn = defaultdict(int)
        anomaly_type_count = defaultdict(int)

        for resource, data in all_results.items():
            anomaly_types = data["anomaly_types"]
            detected_rows = data[method_key]
            detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())

            for row_id, anomaly_type in anomaly_types.items():
                anomaly_type_count[anomaly_type] += 1
                if row_id in detected_set:
                    anomaly_type_tp[anomaly_type] += 1
                else:
                    anomaly_type_fn[anomaly_type] += 1

        per_anomaly_table = []
        all_anomaly_types = set(anomaly_type_tp.keys()) | set(anomaly_type_fn.keys()) | set(anomaly_type_count.keys())

        for anomaly_type in sorted(all_anomaly_types):
            tp = anomaly_type_tp[anomaly_type]
            fn = anomaly_type_fn[anomaly_type]
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            per_anomaly_table.append([anomaly_type, anomaly_type_count[anomaly_type], tp, fn, f"{recall:.3f}"])

        per_anomaly_headers = ["Anomaly Type", "Count", "TP", "FN", "Recall"]
        print(tabulate.tabulate(per_anomaly_table, headers=per_anomaly_headers))

        if STORE:
            method_file_name = method_key.replace("_detected", "") + "_per_anomaly_type_metrics.xlsx"
            method_output_path = os.path.join(tables_dir, method_file_name)
            pd.DataFrame(per_anomaly_table, columns=per_anomaly_headers).to_excel(method_output_path, index=False)

    # Show raw confusion-matrix counts side by side across methods.
    total_val_rows = sum(len(data["val_anomalous"]) for data in all_results.values())
    confusion_counts = {"TP": {}, "FP": {}, "TN": {}, "FN": {}}
    confusion_headers = ["Count", "Autoencoder", "Isolation Forest", "SVM"]
    confusion_output_path = os.path.join(tables_dir, "combined_confusion_matrix_counts.xlsx")
    def _store_confusion_matrix_counts():
        if not STORE:
            return

        rows = [
            [
                count_name,
                confusion_counts[count_name].get("Autoencoder", 0),
                confusion_counts[count_name].get("Isolation Forest", 0),
                confusion_counts[count_name].get("SVM", 0),
            ]
            for count_name in ["TP", "FP", "TN", "FN"]
        ]
        pd.DataFrame(rows, columns=confusion_headers).to_excel(confusion_output_path, index=False)

    atexit.register(_store_confusion_matrix_counts)

    for method_name, method_key in methods:
        total_tp = 0
        total_fp = 0
        total_fn = 0

        for resource, data in all_results.items():
            anomalous_values = data["anomalous_values"]
            detected_rows = data[method_key]
            detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())

            total_tp += sum(1 for row_id in detected_set if row_id in anomalous_values)
            total_fp += sum(1 for row_id in detected_set if row_id not in anomalous_values)
            total_fn += sum(1 for row_id in anomalous_values.keys() if row_id not in detected_set)

        total_tn = total_val_rows - total_tp - total_fp - total_fn
        confusion_counts["TP"][method_name] = total_tp
        confusion_counts["FP"][method_name] = total_fp
        confusion_counts["TN"][method_name] = total_tn
        confusion_counts["FN"][method_name] = total_fn

    confusion_headers = ["Count", "Autoencoder", "Isolation Forest", "SVM"]
    confusion_rows = [
        [
            count_name,
            confusion_counts[count_name].get("Autoencoder", 0),
            confusion_counts[count_name].get("Isolation Forest", 0),
            confusion_counts[count_name].get("SVM", 0),
        ]
        for count_name in ["TP", "FP", "TN", "FN"]
    ]
    print("\nConfusion Matrices (Combined, Raw Counts)")
    print(tabulate.tabulate(confusion_rows, headers=confusion_headers))

    if STORE:
        combined_output_path = os.path.join(tables_dir, "combined_metrics.xlsx")
        pd.DataFrame(combined_table, columns=combined_headers).to_excel(combined_output_path, index=False)
        
        
    import matplotlib.pyplot as plt

    # Extract F1 scores and column counts per resource
    resources_list = []
    f1_scores = []
    column_counts = []

    for resource, data in all_results.items():
        resources_list.append(resource)
        column_counts.append(len(data["column_names"]))
        
        # Calculate F1 score for autoencoder for this resource
        anomalous_values = data["anomalous_values"]
        detected_rows = data["ae_detected"]
        detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())
        
        tp = sum(1 for row_id in detected_set if row_id in anomalous_values)
        fp = sum(1 for row_id in detected_set if row_id not in anomalous_values)
        fn = sum(1 for row_id in anomalous_values.keys() if row_id not in detected_set)
        
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
        f1_scores.append(f1)

    # Create scatter plot
    plt.figure(figsize=(10, 6))
    plt.scatter(column_counts, f1_scores, s=100, alpha=0.6)
    correlation1 = np.corrcoef(column_counts, f1_scores)[0, 1]
    plt.ylabel("F1 Score (Autoencoder)")
    plt.xlabel("Number of Columns")
    plt.title(f"F1 Score vs Number of Columns per Resource (Correlation: {correlation1:.3f})")
    plt.grid(True, alpha=0.3)

    # Add resource labels to points
    for i, resource in enumerate(resources_list):
        plt.annotate(resource, (column_counts[i], f1_scores[i]), fontsize=8, alpha=0.7)

    plt.tight_layout()
    plt.show()
    
    # Create scatter plot for F1 Score vs Training Set Size
    plt.figure(figsize=(10, 6))
    train_rows = [len(data["train_scaled"]) for data in all_results.values()]
    correlation2 = np.corrcoef(train_rows, f1_scores)[0, 1]
    plt.scatter(train_rows, f1_scores, s=100, alpha=0.6, color='orange')
    plt.ylabel("F1 Score (Autoencoder)")
    plt.xlabel("Number of Training Rows")
    plt.title(f"F1 Score vs Training Set Size per Resource (Correlation: {correlation2:.3f})")
    plt.grid(True, alpha=0.3)

    for i, resource in enumerate(resources_list):
        plt.annotate(resource, (train_rows[i], f1_scores[i]), fontsize=8, alpha=0.7)

    plt.tight_layout()
    plt.show()
    
    # Create scatter plot for F1 Score vs Number of Events
    plt.figure(figsize=(10, 6))
    event_counts = []
    for resource, data in all_results.items():
        column_names = data["column_names"]
        num_events = sum(1 for col in column_names if "event" in str(col).lower()) + 1
        event_counts.append(num_events)

    correlation3 = np.corrcoef(event_counts, f1_scores)[0, 1]
    plt.scatter(event_counts, f1_scores, s=100, alpha=0.6, color='green')
    plt.ylabel("F1 Score (Autoencoder)")
    plt.xlabel("Number of Events")
    plt.title(f"F1 Score vs Number of Events per Resource (Correlation: {correlation3:.3f})")
    plt.grid(True, alpha=0.3)

    for i, resource in enumerate(resources_list):
        plt.annotate(resource, (event_counts[i], f1_scores[i]), fontsize=8, alpha=0.7)

    plt.tight_layout()
    plt.show()
    

    # Calculate autoencoder metrics per event type across all resources.
    event_row_count = defaultdict(int)
    event_tp = defaultdict(int)
    event_fp = defaultdict(int)
    event_fn = defaultdict(int)

    for resource, data in all_results.items():
        val_rows = data["val_anomalous"]
        column_names = data["column_names"]
        anomalous_values = data["anomalous_values"]
        detected_rows = data["ae_detected"]

        event_columns = extract_event_column_indices(column_names)
        detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())

        for row_id in range(len(val_rows)):
            event_name = get_event_for_row(val_rows[row_id], event_columns)
            event_row_count[event_name] += 1
            is_actual_anomaly = row_id in anomalous_values
            is_detected_anomaly = row_id in detected_set

            if is_actual_anomaly and is_detected_anomaly:
                event_tp[event_name] += 1
            elif not is_actual_anomaly and is_detected_anomaly:
                event_fp[event_name] += 1
            elif is_actual_anomaly and not is_detected_anomaly:
                event_fn[event_name] += 1

    print(f"\n{'='*60}")
    print("Autoencoder Results Per Event Type")
    print(f"{'='*60}\n")

    per_event_table = []
    all_events = set(event_tp.keys()) | set(event_fp.keys()) | set(event_fn.keys())
    for event_name in all_events:
        tp = event_tp[event_name]
        fp = event_fp[event_name]
        fn = event_fn[event_name]

        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

        per_event_table.append([event_name, event_row_count[event_name], tp, fp, fn, precision, recall, f1])

    per_event_table.sort(key=lambda row: row[7], reverse=True)
    per_event_table_for_print = [
        [row[0], row[1], row[2], row[3], row[4], f"{row[5]:.3f}", f"{row[6]:.3f}", f"{row[7]:.3f}"]
        for row in per_event_table
    ]

    per_event_headers = ["Event", "Val Rows", "TP", "FP", "FN", "Precision", "Recall", "F1-Score"]
    per_event_table_text = tabulate.tabulate(per_event_table_for_print, headers=per_event_headers)

    # Not printed because of the size

    if STORE:
        per_event_output_path = os.path.join(tables_dir, "autoencoder_per_event_metrics.xlsx")
        pd.DataFrame(per_event_table_for_print, columns=per_event_headers).to_excel(per_event_output_path, index=False)

        print(f"\nStored result tables in '{tables_dir}'")
        

    # Analyze reconstruction examples for one fixed resource (ov_1):
    # - one true positive per anomaly type
    # - one false negative
    # - one false positive
    print(f"\n{'='*60}")
    print("Autoencoder Reconstruction Examples")
    print(f"{'='*60}\n")

    example_resource = "ov_1"
    if example_resource not in all_results:
        print(f"Resource '{example_resource}' not found in selected resources. Skipping reconstruction examples.")
    else:
        data = all_results[example_resource]

        all_anomaly_types = sorted(set(data["anomaly_types"].values()))
        examples = {
            "TP_BY_TYPE": {anomaly_type: None for anomaly_type in all_anomaly_types},
            "TN": None,
            "FN": None,
            "FP": None,
        }

        def _build_example(row_id, resource_data):
            val_rows_reversed = resource_data["_val_rows_reversed"]
            reconstructions_reversed = resource_data["_reconstructions_reversed"]
            val_rows_scaled = resource_data["_val_rows_scaled"]
            reconstructions_scaled = resource_data["_reconstructions_scaled"]

            is_actual_anomaly = row_id in resource_data["anomalous_values"]
            clean_original_reversed = resource_data["_val_scaled_clean_reversed"][row_id] if is_actual_anomaly else None

            row_loss = float(np.mean((val_rows_scaled[row_id] - reconstructions_scaled[row_id]) ** 2))
            per_column_loss = (val_rows_scaled[row_id] - reconstructions_scaled[row_id]) ** 2

            return {
                "resource": example_resource,
                "row_id": row_id,
                "original": val_rows_reversed[row_id],  # input row (possibly anomalous)
                "clean_original": clean_original_reversed,  # from val_scaled (only for actual anomalies)
                "reconstruction": reconstructions_reversed[row_id],
                "column_names": resource_data["column_names"],
                "anomaly_columns": list(resource_data["anomalous_values"].get(row_id, [])),
                "anomaly_type": resource_data["anomaly_types"].get(row_id, "UNKNOWN_TYPE"),
                "row_loss": row_loss,
                "per_column_loss": per_column_loss,
                "is_actual_anomaly": is_actual_anomaly,
            }

        # Precompute arrays only for ov_1
        resource_scaling = data["scaling"]
        data["_val_rows_scaled"] = to_numpy_2d(data["val_anomalous"])
        data["_reconstructions_scaled"] = to_numpy_2d(data["ae_reconstructions"])
        data["_val_rows_reversed"] = resource_scaling.inverse_transform(data["_val_rows_scaled"])
        data["_reconstructions_reversed"] = resource_scaling.inverse_transform(data["_reconstructions_scaled"])
        train_rows_scaled = to_numpy_2d(data["train_scaled"])
        train_rows_reversed = resource_scaling.inverse_transform(train_rows_scaled)

        # StandardScaler exposes per-feature std via scale_.
        scaling_std = getattr(resource_scaling, "scale_", None)
        if scaling_std is None:
            scaling_std = np.std(train_rows_reversed, axis=0)

        # Median is not part of StandardScaler, so derive it from unscaled training data.
        scaling_median = np.median(train_rows_reversed, axis=0)
        scaling_stats = {
            "feature_names": [str(c) for c in data["column_names"]],
            "std_per_feature": [float(v) for v in np.asarray(scaling_std)],
            "median_per_feature": [float(v) for v in np.asarray(scaling_median)],
        }

        # Load clean validation rows (val_scaled) for "original data" comparison
        _, _, val_scaled_clean, _, _ = read_and_prepare_data(example_resource, load_preprepared=reread_prepared_data, prints=False)
        data["_val_scaled_clean_reversed"] = resource_scaling.inverse_transform(to_numpy_2d(val_scaled_clean))

        val_rows = data["val_anomalous"]
        anomalous_values = data["anomalous_values"]
        anomaly_types = data["anomaly_types"]
        detected_rows = data["ae_detected"]
        detected_set = set(detected_rows) if not isinstance(detected_rows, dict) else set(detected_rows.keys())

        # Seeded random selection of examples (reproducible across runs)
        rng = random.Random(RANDOM_SEED)

        tp_candidates_by_type = defaultdict(list)
        fn_candidates = []
        fp_candidates = []
        tn_candidates = []

        for row_id in range(len(val_rows)):
            is_actual_anomaly = row_id in anomalous_values
            is_detected_anomaly = row_id in detected_set

            if is_actual_anomaly and is_detected_anomaly:
                anomaly_type = anomaly_types.get(row_id, "UNKNOWN_TYPE")
                tp_candidates_by_type[anomaly_type].append(row_id)
            elif is_actual_anomaly and not is_detected_anomaly:
                fn_candidates.append(row_id)
            elif not is_actual_anomaly and is_detected_anomaly:
                fp_candidates.append(row_id)
            else:
                tn_candidates.append(row_id)

        def _normalize_row_id(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                return value

        for anomaly_type in all_anomaly_types:
            candidates = tp_candidates_by_type.get(anomaly_type, [])
            if candidates:
                chosen_row = rng.choice(candidates)
                examples["TP_BY_TYPE"][anomaly_type] = _build_example(chosen_row, data)

        if tn_candidates:
            examples["TN"] = _build_example(rng.choice(tn_candidates), data)

        if fn_candidates:
            examples["FN"] = _build_example(rng.choice(fn_candidates), data)

        if fp_candidates:
            examples["FP"] = _build_example(rng.choice(fp_candidates), data)

        def _print_example(title, example):
            print(f"\n{title}")
            if example is None:
                print("  No example found.")
                return

            print(f"  Resource: {example['resource']}, Row ID: {example['row_id']}, Anomaly Type: {example['anomaly_type']}")
            print(f"  AE Loss (row MSE on scaled input vs reconstruction): {example['row_loss']:.6f}")

            anomaly_col_names = [example["column_names"][col_idx] for col_idx in example["anomaly_columns"]]
            if anomaly_col_names:
                print(f"  Anomaly columns: {', '.join(map(str, anomaly_col_names))}")

            anomaly_col_set = set(example["anomaly_columns"])
            comparison = []
            for col_idx, col_name in enumerate(example["column_names"]):
                input_value = example["original"][col_idx]
                reconstruction = example["reconstruction"][col_idx]
                clean_value = example["clean_original"][col_idx] if example["clean_original"] is not None else None
                column_loss = float(example["per_column_loss"][col_idx])
                is_anomaly = " ⚠" if col_idx in anomaly_col_set else ""

                comparison.append([
                    f"{col_name}{is_anomaly}",
                    f"{input_value:.4f}",
                    f"{clean_value:.4f}" if clean_value is not None else "-",
                    f"{reconstruction:.4f}",
                    f"{column_loss:.6f}",
                ])

            print(tabulate.tabulate(comparison, headers=["Column", "Input (Val Anomalous)", "Original (val_scaled)", "Reconstruction", "Loss (MSE)"]))

        def _slug(text):
            return re.sub(r"[^a-zA-Z0-9_.-]+", "_", str(text)).strip("_") or "unknown"

        def _to_serializable(value):
            if isinstance(value, np.ndarray):
                return value.tolist()
            if isinstance(value, (np.integer,)):
                return int(value)
            if isinstance(value, (np.floating,)):
                return float(value)
            if isinstance(value, dict):
                return {str(k): _to_serializable(v) for k, v in value.items()}
            if isinstance(value, (list, tuple)):
                return [_to_serializable(v) for v in value]
            return value

        examples_dir = os.path.join(tables_dir, "reconstruction_examples")
        if STORE:
            os.makedirs(examples_dir, exist_ok=True)

        def _save_example(example_key, example):
            if not STORE or example is None:
                return

            anomaly_col_names = [str(example["column_names"][col_idx]) for col_idx in example["anomaly_columns"]]
            per_column = []
            anomaly_col_set = set(example["anomaly_columns"])

            for col_idx, col_name in enumerate(example["column_names"]):
                per_column.append({
                    "column_index": col_idx,
                    "column_name": str(col_name),
                    "is_anomaly_column": col_idx in anomaly_col_set,
                    "input_val_anomalous": float(example["original"][col_idx]),
                    "original_val_scaled": float(example["clean_original"][col_idx]) if example["clean_original"] is not None else None,
                    "reconstruction": float(example["reconstruction"][col_idx]),
                    "loss_mse": float(example["per_column_loss"][col_idx]),
                })

            payload = {
                "example_type": example_key,  # TP_<type>, FN, FP
                "resource": example["resource"],
                "row_id": int(example["row_id"]),
                "anomaly_type": str(example["anomaly_type"]),
                "is_actual_anomaly": bool(example["is_actual_anomaly"]),
                "row_loss_mse_scaled": float(example["row_loss"]),
                "autoencoder_threshold": float(data["ae_threshold"]),
                "z_normalization_scaling": scaling_stats,
                "anomaly_columns_indices": [int(i) for i in example["anomaly_columns"]],
                "anomaly_columns_names": anomaly_col_names,
                "column_names": [str(c) for c in example["column_names"]],
                "input_row_val_anomalous": _to_serializable(example["original"]),
                "original_row_val_scaled_clean": _to_serializable(example["clean_original"]) if example["clean_original"] is not None else None,
                "reconstruction_row": _to_serializable(example["reconstruction"]),
                "per_column_details": per_column,
            }

            file_name = f"{_slug(example_key)}__{_slug(example['resource'])}__row_{int(example['row_id'])}.json"
            output_path = os.path.join(examples_dir, file_name)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False)

            # Also store a thesis-ready XLSX version of the same example.
            xlsx_file_name = f"{_slug(example_key)}__{_slug(example['resource'])}__row_{int(example['row_id'])}.xlsx"
            xlsx_output_path = os.path.join(examples_dir, xlsx_file_name)

            wb = Workbook()
            ws = wb.active
            ws.title = "Example"

            title = f"Reconstruction Example: {example_key}"
            ws.merge_cells("A1:E1")
            ws["A1"] = title
            ws["A1"].font = Font(name="Calibri", size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A2:E2")
            ws["A2"] = (
                f"Resource: {example['resource']} | Row ID: {int(example['row_id'])} | "
                f"Anomaly Type: {example['anomaly_type']} | Row Loss (MSE): {float(example['row_loss']):.6f}"
            )
            ws["A2"].font = Font(name="Calibri", size=11, italic=True)
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

            header_row = 4
            headers = ["Sensor Name", "Input", "Reconstruction", "Original Data", "Loss (MSE)"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            anomaly_col_set = set(example["anomaly_columns"])
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            for row_offset, col_name in enumerate(example["column_names"], start=1):
                row_idx = header_row + row_offset
                col_idx = row_offset - 1

                input_value = float(example["original"][col_idx])
                reconstruction_value = float(example["reconstruction"][col_idx])
                original_data_value = (
                    float(example["clean_original"][col_idx])
                    if example["clean_original"] is not None
                    else None
                )
                value_loss = float(example["per_column_loss"][col_idx])

                ws.cell(row=row_idx, column=1, value=str(col_name))
                ws.cell(row=row_idx, column=2, value=input_value)
                ws.cell(row=row_idx, column=3, value=reconstruction_value)
                ws.cell(row=row_idx, column=4, value=original_data_value)
                ws.cell(row=row_idx, column=5, value=value_loss)

                if col_idx in anomaly_col_set:
                    for col in range(1, 6):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                        )

                for col in range(1, 6):
                    c = ws.cell(row=row_idx, column=col)
                    c.border = thin_border
                    c.alignment = Alignment(vertical="center")

            for col in range(2, 6):
                for row in range(header_row + 1, header_row + 1 + len(example["column_names"])):
                    ws.cell(row=row, column=col).number_format = "0.0000"

            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 16
            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A{header_row}:E{header_row + len(example['column_names'])}"
            ws.sheet_view.showGridLines = False

            wb.save(xlsx_output_path)

        # Print and save one TP per anomaly type
        for anomaly_type in all_anomaly_types:
            key = f"TP_{anomaly_type}"
            example = examples["TP_BY_TYPE"][anomaly_type]
            _print_example(f"TP Example - {anomaly_type}", example)
            _save_example(key, example)

        _print_example("TN Example", examples["TN"])
        _save_example("TN", examples["TN"])

        # Print and save one FN and one FP
        _print_example("FN Example", examples["FN"])
        _save_example("FN", examples["FN"])

        _print_example("FP Example", examples["FP"])
        _save_example("FP", examples["FP"])

        if STORE:
            print(f"\nStored reconstruction examples in '{examples_dir}'")

        # Cleanup temporary cached arrays
        data.pop("_val_rows_scaled", None)
        data.pop("_reconstructions_scaled", None)
        data.pop("_val_rows_reversed", None)
        data.pop("_reconstructions_reversed", None)
        data.pop("_val_scaled_clean_reversed", None)
            
            
    # Load and display hyperparameters from the hyperparameters folder
    hyperparameters_dir = "hyperparameters"
    hyperparams_table = []

    if os.path.exists(hyperparameters_dir):
        for resource in all_resources:
            param_file = os.path.join(hyperparameters_dir, f"{resource}_hyperparameters.npy")
            if os.path.exists(param_file):
                params = np.load(param_file, allow_pickle=True).item()
                hyperparams_table.append([resource, params])
        
        print(f"\n{'='*60}")
        print("Hyperparameters per Resource")
        print(f"{'='*60}\n")
        
        if hyperparams_table:
            # Extract all unique parameter keys
            all_keys = set()
            for _, params in hyperparams_table:
                all_keys.update(params.keys())
            all_keys = sorted(list(all_keys))
            
            # Build table with resources as rows and parameters as columns
            headers = ["Resource"] + all_keys
            table_rows = []
            for resource, params in hyperparams_table:
                row = [resource]
                for key in all_keys:
                    row.append(params.get(key, "N/A"))
                table_rows.append(row)
            
            print(tabulate.tabulate(table_rows, headers=headers))
            
            if STORE:
                hyperparams_output_path = os.path.join(tables_dir, "hyperparameters.xlsx")
                pd.DataFrame(table_rows, columns=headers).to_excel(hyperparams_output_path, index=False)
